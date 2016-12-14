import requests
## This script should:
import requests
import shutil
import flattentool
import json
import collections
import openpyxl
from openpyxl import load_workbook


# 1. Download the excel file below
# 2. Convert to JSON with flatten tool
# 3. Place in standard/docs/en/examples/ppp/full.json
# 4. Split out each of the releases into it's own file named after the OCID

url = 'https://docs.google.com/spreadsheets/d/1gvzLoImbnWvty7lfe5mx2h3v3cELj2rsJHBhj_Yo9Cs/pub?output=xlsx'
response = requests.get(url, stream=True)
with open('download.xlsx', 'wb+') as out_file:
    shutil.copyfileobj(response.raw, out_file)


## Check for a row with column paths within the first five rows (by checking for / in more than one of the first 10 columns)
## WORK IN PROGRESS
## wb = load_workbook('download.xlsx')
## 
## for sheet in wb.worksheets:
##   for r in range(1,5):
##     rowSlashCount = 0
##     for c in range(1,10):
##       if sheet.cell(row=r, column=c).value and "/" in str(sheet.cell(row=r, column=c).value): 
##         rowSlashCount = rowSlashCount + 1
##     if rowSlashCount > 2:
##       print(sheet[1])
##       sheet[1] = sheet[r]
##       print("Remove rows before" + str(r) + "and escape")
##       break


flattentool.unflatten(
  'download.xlsx',
  output_name='standard/docs/en/examples/ppp/full.json',
  input_format='xlsx',
  root_list_path='releases',
  root_id='ocid',
  encoding='utf-8'
)

with open('standard/docs/en/examples/ppp/full.json') as full_file:
    full = json.load(full_file)

    split_releases = collections.defaultdict(list)

    for release in full['releases']:
        if 'id' not in release:
            print('Warning, release without an id')
            print(json.dumps(release, indent=2))
            continue
        split_releases[release['id']].append(release)

    for key, value in split_releases.items():
        with open('standard/docs/en/examples/ppp/{}.json'.format(key), 'w+') as release_file:
            json.dump({"releases": value}, release_file, indent=2)
